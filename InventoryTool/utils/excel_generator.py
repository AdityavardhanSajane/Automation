import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

def generate_excel_file(server_data, filename):
    """
    Generate an Excel file with server inventory data
    
    Args:
        server_data: List of dictionaries with component and server information
        filename: Name of the Excel file to create
    
    Returns:
        Path to the generated Excel file
    """
    try:
        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Server Inventory"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0366D6", end_color="0366D6", fill_type="solid")
        component_font = Font(bold=True)
        component_fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center')
        
        # Add headers
        headers = [
            "Component", "Group Name", "Server Name", "Environment", 
            "OS Information", "Enabled"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Freeze the header row
        ws.freeze_panes = 'A2'
        
        # Add data
        row_num = 2
        for component_data in server_data:
            component_name = component_data['component']
            servers = component_data['servers']
            
            if not servers:
                # Add the component row even if there are no servers
                ws.cell(row=row_num, column=1).value = component_name
                ws.cell(row=row_num, column=1).font = component_font
                ws.cell(row=row_num, column=1).fill = component_fill
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).border = border
                
                row_num += 1
                continue
            
            for server in servers:
                ws.cell(row=row_num, column=1).value = component_name
                ws.cell(row=row_num, column=2).value = server.get('group_name', '')
                ws.cell(row=row_num, column=3).value = server.get('server_name', '')
                ws.cell(row=row_num, column=4).value = server.get('environment', '')
                ws.cell(row=row_num, column=5).value = server.get('os_info', '')
                ws.cell(row=row_num, column=6).value = 'Yes' if server.get('enabled', False) else 'No'
                
                # Apply styles
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).border = border
                
                # Apply component style to first cell
                ws.cell(row=row_num, column=1).font = component_font
                ws.cell(row=row_num, column=1).fill = component_fill
                
                row_num += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the file to the current directory
        file_path = os.path.join(os.getcwd(), filename)
        wb.save(file_path)
        logger.info(f"Excel file generated: {file_path}")
        
        return file_path
    
    except Exception as e:
        logger.error(f"Error generating Excel file: {str(e)}")
        raise

